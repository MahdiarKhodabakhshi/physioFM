#!/usr/bin/env python3
"""Build docs/paper_results_tables.xlsx — the discussion workbook for the paper.

Sheets:
  0. Numbers       — clean, numbers-only tables for the meeting (no prose in cells).
  1. Method        — what the final method IS (architecture, tokens, training, evaluation), naming key,
                     and the open method decisions.
  2. Main results  — the results we would put in the paper, one row per task x setting, with the
                     final-method column, the matched controls, baselines, SOTA, seeds, sources.
  3. Ablations     — every control / ablation, one row each: question, arms + numbers, delta, verdict,
                     seeds, source, include?
  4. Do not claim  — retracted or unsupported statements, with the evidence that killed them.
Every number traces to docs/experiments/EXP-* and results/ (source column).
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parents[1] / "docs" / "paper_results_tables.xlsx"

HDR_FILL = PatternFill("solid", fgColor="17203A")
HDR_FONT = Font(bold=True, color="FFFFFF")
SEC_FILL = PatternFill("solid", fgColor="E3F2F1")
SEC_FONT = Font(bold=True, color="0E5C5B")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="D5DAE4")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
INCLUDE_FILL = {"yes": PatternFill("solid", fgColor="E3F2F1"),
                "discuss": PatternFill("solid", fgColor="FBEFDF"),
                "no": PatternFill("solid", fgColor="F9E5E5"),
                "appendix": PatternFill("solid", fgColor="EEF1F6")}


def write_sheet(ws, headers, rows, widths, include_col=None, section_col=0):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = WRAP; cell.border = BORDER
    for r in rows:
        if isinstance(r, str):  # section row
            ws.append([r])
            row = ws.max_row
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
            cell = ws.cell(row=row, column=1); cell.fill = SEC_FILL; cell.font = SEC_FONT
            continue
        ws.append(r)
        row = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=c); cell.alignment = WRAP; cell.border = BORDER
        if include_col is not None:
            v = str(ws.cell(row=row, column=include_col + 1).value or "").lower()
            for k, f in INCLUDE_FILL.items():
                if v.startswith(k):
                    ws.cell(row=row, column=include_col + 1).fill = f
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


TITLE_FONT = Font(bold=True, size=12, color="17203A")
NUM_ALIGN = Alignment(horizontal="right", vertical="center")


FINAL_FILL = PatternFill("solid", fgColor="C8EBE7")
FINAL_FONT = Font(bold=True, color="0E5C5B")
CONTROL_FILL = PatternFill("solid", fgColor="EEF1F6")
# (block-title prefix, row label, column header) -> "final" | "control"
FINAL_CELLS = {
    ("Sleep-EDF-78 — subject-disjoint", "tf64 (2×64)", "Pretrained"): "final",
    ("Sleep-EDF-78 — subject-disjoint", "tf64 (2×64)", "No pretrain"): "control",
    ("Sleep-EDF-78 — Cohen", "tf64", "Pretrained"): "final",
    ("Sleep-EDF-78 — Cohen", "tf64", "No pretrain"): "control",
    ("CHB-MIT", "Fine-tuned — pretrained", "bal-acc %"): "final",
    ("CHB-MIT", "Fine-tuned — pretrained", "ROC-AUC"): "final",
    ("CHB-MIT", "Fine-tuned — no pretrain", "bal-acc %"): "control",
    ("CHB-MIT", "Fine-tuned — no pretrain", "ROC-AUC"): "control",
    ("Streaming", "Causal — pretrained", "Offline (whole window)"): "final",
    ("Streaming", "Causal — pretrained", "Online (past only)"): "final",
    ("Streaming", "Causal — no pretrain", "Offline (whole window)"): "control",
    ("Streaming", "Causal — no pretrain", "Online (past only)"): "control",
}


def write_blocks(ws, blocks, widths):
    """Clean numbers-only sheet: a list of (title, headers, rows[, note]) blocks separated by a blank row."""
    ws.cell(row=1, column=1, value="FINAL METHOD = green cells: structured tokens (tf64 on sleep; DE on CHB-MIT — tf64 not run there) + causal decoder + input-space PC pretraining, fine-tuned end-to-end. Grey = its matched no-pretrain control (same architecture, random init).").font = Font(italic=True, color="0E5C5B")
    r = 3
    for block in blocks:
        title, headers, rows = block[0], block[1], block[2]
        note = block[3] if len(block) > 3 else None
        ws.cell(row=r, column=1, value=title).font = TITLE_FONT
        r += 1
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.border = BORDER
            cell.alignment = Alignment(horizontal="left" if c == 1 else "right", vertical="center", wrap_text=True)
        r += 1
        for row in rows:
            for c, v in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.border = BORDER
                cell.alignment = Alignment(horizontal="left", vertical="center") if c == 1 else NUM_ALIGN
                if isinstance(v, float) and abs(v) >= 10:
                    cell.number_format = "0.0"   # accuracies; deltas / AUC / kappa keep their own precision
                for (tp, rl, ch), kind in FINAL_CELLS.items():
                    if title.startswith(tp) and row[0] == rl and headers[c - 1] == ch:
                        cell.fill = FINAL_FILL if kind == "final" else CONTROL_FILL
                        if kind == "final":
                            cell.font = FINAL_FONT
            r += 1
        if note:
            ws.cell(row=r, column=1, value=note).font = Font(italic=True, color="5B6579", size=9)
            r += 1
        r += 1
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def numbers_sheet(ws):
    blocks = [
        ("Sleep-EDF-78 — subject-disjoint 5-fold — accuracy %, FINE-TUNED",
         ["Input tokens", "Pretrained", "No pretrain", "Δ pretrain", "Raw features → logreg", "HGB", "SOTA (unverified)", "Pretrain seeds"],
         [["tf64 (2×64)", 77.9, 77.0, 0.85, 72.8, 75.2, "81–85", 4],
          ["DE (2×5)", 75.5, 73.1, 2.5, 67.9, 69.6, "81–85", 4],
          ["raw 200 ms (2×20)", 75.5, 74.2, 1.2, None, None, "81–85", 1],
          ["raw per-electrode (1×20)", 76.4, 74.6, 1.8, None, None, "81–85", 1]],
         "± over seeds: tf64 0.4 / 0.2; DE 0.3 / 0.5. κ: tf64 .71 / .69; DE .672 / .645; raw .676 / .658."),
        ("Sleep-EDF-78 — Cohen's κ, FINE-TUNED",
         ["Input tokens", "Pretrained", "No pretrain", "Raw features → logreg", "SOTA (unverified)"],
         [["tf64", 0.71, 0.69, 0.64, "0.77–0.83"],
          ["DE", 0.672, 0.645, 0.575, "0.77–0.83"],
          ["raw 200 ms", 0.676, 0.658, None, "0.77–0.83"]]),
        ("CHB-MIT seizure detection — leave-one-patient-out (24 patients)",
         ["Arm", "bal-acc %", "ROC-AUC", "Seeds"],
         [["Fine-tuned — pretrained", 78.4, 0.863, 1],
          ["Fine-tuned — pretrained (Jul, 3-seed mean)", 78.2, 0.867, 3],
          ["Fine-tuned — no pretrain", 80.2, 0.874, 1],
          ["Fine-tuned — no pretrain (Jul, 3-seed mean)", 79.1, 0.860, 3],
          ["Fine-tuned — latent objective", 79.7, 0.879, 1],
          ["Frozen — pretrained", 77.4, 0.852, 1],
          ["Frozen — no pretrain", 67.5, 0.741, 1],
          ["Raw DE → logreg", 72.4, 0.806, None],
          ["Raw DE → HGB", 74.0, 0.851, None],
          ["SOTA cross-patient (unverified)", None, "0.91–0.99", None]],
         "Per-patient sd ±12–16 bal-acc: the three fine-tuned arms are indistinguishable. tf64 gives no headroom on this task (linear 72.4 = 72.4)."),
        ("Streaming evaluation — Sleep-EDF-78, DE tokens, FINE-TUNED (accuracy %)",
         ["Arm", "Offline (whole window)", "Online (past only)", "Tokens per decision"],
         [["Causal — no pretrain", 73.2, 73.2, 1],
          ["Bidirectional twin — no pretrain", 74.1, 70.4, 190],
          ["Causal — latent", 74.1, 74.1, 1],
          ["Causal — pretrained", 75.4, 75.4, 1]],
         "1 seed, 5 folds. Causal beats bidirectional online by +2.8 (no pretrain) / +5.0 (pretrained)."),
        ("Frozen-probe settings (never fine-tuned) — accuracy % unless noted",
         ["Task", "Pretrained", "No pretrain", "Raw features → logreg", "Seeds"],
         [["Emotion SEED-IV, smoothed", 59.3, 56.9, 62.8, 3],
          ["Emotion SEED-IV, un-smoothed", 51.7, 40.7, 55.3, 3],
          ["Motor imagery BCI-IV-2a", 41.7, 43.5, 51.1, 3],
          ["Seizure prediction — bal-acc", 66.8, 71.3, 65.2, 1],
          ["Seizure prediction — AUC", 0.769, 0.793, 0.710, 1]]),
        ("Ablation — pretraining Δ (pretrained − no pretrain): frozen probe vs fine-tuned",
         ["Setting", "Frozen Δ", "Fine-tuned Δ", "Inflation ×"],
         [["Sleep DE", 9.8, 2.5, 3.9],
          ["Sleep tf64", 11.2, 0.85, 13.2],
          ["Sleep raw", 14.7, 1.2, 12.3],
          ["Seizure DE", 8.1, 0.0, None]]),
        ("Ablation — pretraining objective (Sleep DE, fine-tuned accuracy %, seed 42)",
         ["Objective", "Accuracy"],
         [["Input-space PC (next 16 tokens, MSE)", 75.4],
          ["Latent targets (JEPA/data2vec)", 74.1],
          ["Latent — delta targets", 73.9],
          ["Latent — cosine, no normalisation", 72.7],
          ["Latent — variance term", 73.2],
          ["Latent — EMA 0.99", 74.0],
          ["Latent — p_out 4", 74.0],
          ["No pretraining", 73.2]],
         "4-seed means: input-PC 75.5, latent 73.7, no pretrain 73.1."),
        ("Ablation — linear saturation of the input features",
         ["Features", "Logistic regression", "MLP", "HGB", "Headroom (best − linear)"],
         [["Sleep DE (acc)", 67.9, 67.2, 69.6, 1.7],
          ["Sleep tf64 (acc)", 72.8, 71.9, 75.2, 2.3],
          ["Seizure DE (bal-acc)", 72.4, 69.5, 74.0, 1.6],
          ["Seizure tf64 (bal-acc)", 72.4, 69.9, 72.6, 0.2],
          ["Seizure DE (AUC)", 0.807, 0.824, 0.851, 0.044],
          ["Seizure tf64 (AUC)", 0.809, 0.824, 0.849, 0.040]]),
        ("Ablation — dimension-matched control (frozen probe, accuracy %)",
         ["Setting", "Raw features", "Random 256-d projection", "Pretrained encoder", "Δ vs projection", "Concat raw ‖ encoder"],
         [["Sleep DE", 67.9, 69.3, 72.6, 3.3, 73.3],
          ["Sleep tf64", 72.8, 73.3, 76.8, 3.5, 77.4],
          ["Emotion SEED-IV smoothed", 62.8, 60.2, 60.0, -0.2, 62.5],
          ["Motor imagery", 51.1, 49.9, 40.6, -9.3, 43.8]]),
        ("Ablation — structured vs per-electrode raw tokens (Sleep, fine-tuned accuracy %, 1 seed)",
         ["Arm", "Structured (2×20)", "Per-electrode (1×20)"],
         [["Pretrained", 75.5, 76.4], ["Latent", 74.7, 75.5], ["No pretrain", 74.2, 74.6]]),
        ("Ablation — label fraction (Sleep DE, Δ pretrained − no pretrain, accuracy points)",
         ["Training labels", "Frozen Δ", "Fine-tuned Δ"],
         [["1 %", 12.7, 1.9], ["5 %", 11.8, 2.4], ["10 %", 10.8, 2.4], ["100 %", 9.8, 2.2]]),
        ("Reference — Phase 1 vs Phase 2 (emotion, frozen probe, accuracy %)",
         ["Setting", "SEED-V", "SEED-IV", "Chance"],
         [["Univariate TimesFM (Phase 1)", "23–25", "26–28", "20 / 25"],
          ["Structured patch, no pretrain", 48.6, 60.7, "20 / 25"],
          ["Structured patch, pretrained", 45.6, 57.5, "20 / 25"],
          ["Raw DE → logreg", 51.4, 62.8, "20 / 25"],
          ["Raw DE + per-series instance norm", 18.0, 26.8, "20 / 25"]]),
        ("Reference — PC-SSL leakage audit (accuracy %)",
         ["Dataset", "Published", "Same code, random window split", "Same code, trial-disjoint split", "Raw DE → logreg"],
         [["SEED-V", 92.4, 65.8, 39.8, 51.4], ["SEED-IV", 84.5, 70.4, 44.7, 62.8]]),
    ]
    write_blocks(ws, blocks, [42, 16, 16, 16, 20, 12, 18, 14])


def main() -> None:
    wb = Workbook()

    # ------------------------------------------------------------------ 0. Numbers (clean)
    ws = wb.active; ws.title = "Numbers"
    numbers_sheet(ws)

    # ------------------------------------------------------------------ 1. Method
    ws = wb.create_sheet("Method")
    headers = ["Item", "Final method (proposed for the paper)", "Alternatives we measured", "Decision / note"]
    rows = [
        "A. What the model is",
        ["Architecture", "Structured-token causal decoder transformer: TimesFM-2.5 decoder layers (RoPE, SDPA), d=256, 6 layers, 8 heads, ~2.4M params; linear token embedder; per-epoch classification head.",
         "Frozen TimesFM-2.5 stack (E1b), frozen random 1280x20 stack (F3): neither beats the small from-scratch stack.", "Fixed."],
        ["Token", "One token = one labelled epoch; the whole (channels x features) matrix, flattened. FINAL: tf64 = 64 log-spaced Welch log-power bins per channel (sleep 2x64=128-d).",
         "DE (5 bands/channel; the proposal's token) -> 75.5 sleep; raw 200 ms tokens (2x20 samples, 150 tokens/epoch) -> 75.5; per-electrode raw -> 76.4.",
         "DISCUSS: tf64 is +2.3 over DE on sleep but gives no headroom on seizure (tf64 ladder on CHB-MIT skipped). Paper can present DE as the proposal's token and tf64 as the improved front end."],
        ["Normalisation", "Fixed per-(channel,feature) corpus standardisation; NO per-series instance norm (RevIN).",
         "Instance norm collapses raw DE to chance (51.4->18.0 SEED-V, 62.8->26.8 SEED-IV).", "Fixed (Phase-1 lesson)."],
        ["Causality", "Causal (each epoch sees only the past) -> streaming-capable, 1 token per decision with a KV cache.",
         "Bidirectional twin: +0.9 offline, -2.8/-5.0 online.", "Fixed; the streaming result is a selling point."],
        "B. Training",
        ["Pretraining (the proposal's contribution)", "Input-space predictive coding: predict the next 16 tokens (multi-step MSE), 60 epochs (DE/tf64) / 10 epochs (raw), single dataset, no labels. Worth +2.5 (DE) / +0.85 (tf64) / +1.2 (raw) on sleep, ~0 on seizure, fine-tuned.",
         "Latent-target (JEPA/data2vec) objective and 5 variants: worse than input-PC (73.7 vs 75.5 sleep); masked reconstruction (F9): same null on emotion. Random init: within 0.85-2.5 points.",
         "DISCUSS: keep pretraining as part of 'the method' (small positive on sleep) or present the model as trained from scratch with pretraining as an ablation? Recommendation: report both columns everywhere; call the final method 'pretrained' only if we keep the sleep +0.85..+2.5 as a claim."],
        ["Downstream training", "End-to-end fine-tuning (encoder + head), class-weighted CE, AdamW 1e-4, nights/files chunked (400 epochs DE/tf64; 20 epochs raw), 8 epochs (sleep DE/tf64), 3 (raw), 4 (seizure).",
         "Frozen encoder + logistic regression (the July protocol): inflates pretraining 3-12x.", "Fixed: fine-tuned numbers only in the paper; frozen shown once as the inflation example."],
        "C. Evaluation",
        ["Sleep", "Sleep-EDF Cassette, 153 nights / 78 subjects / 195k epochs, Fpz-Cz + Pz-Oz, 5 classes (W/N1/N2/N3/REM), 30-min wake trimming, subject-disjoint 5-fold (seed 42 split); acc / macro-F1 / kappa; pretraining seeds 42,1,2,3.", "", "Fixed."],
        ["Seizure", "CHB-MIT, 24 patients / 682 files / 1.76M 2-s epochs (0.32% seizure), 18-ch bipolar core montage, leave-one-patient-out; bal-acc / sens / spec / ROC-AUC, class-weighted; per-patient paired tests.", "", "Fixed."],
        ["Emotion / MI", "SEED-IV (smoothed + un-smoothed), PC-SSL subject-dependent folds; BCI-IV-2a session hold-out. Frozen-probe only (never fine-tuned).", "", "DISCUSS: keep as 'where the encoder does not help' (frozen) or drop/appendix."],
        ["Controls reported with every result", "matched random-init (same architecture, same seed); raw-features->logreg; best non-deep baseline (HGB); dimension-matched random projection (frozen).", "", "Fixed."],
        "D. Naming",
        ["Names in tables", "PhysioFM (pretrained) = physiofm_pc; PhysioFM (no pretrain) = physiofm_rand; PhysioFM (latent) = physiofm_latent; raw features -> logreg = raw_de / raw_tf64.", "", "DISCUSS: 'PhysioFM' implies a foundation model (no joint pretraining, no zero-shot transfer was done). Keep or rename?"],
    ]
    write_sheet(ws, headers, rows, [28, 70, 60, 60])

    # ------------------------------------------------------------------ 2. Main results
    ws = wb.create_sheet("Main results")
    headers = ["Task", "Dataset / protocol", "Input tokens", "Metric", "Final method: PhysioFM (pretrained)", "PhysioFM (no pretrain)", "Pretraining delta",
               "raw features -> logreg", "Best non-deep baseline (HGB)", "Published SOTA (UNVERIFIED web figures)", "Seeds", "Source", "Include in paper?", "Notes for discussion"]
    rows = [
        "Per-epoch tasks, FINE-TUNED (the protocol we report)",
        ["Sleep staging", "Sleep-EDF-78, subject-disjoint 5-fold", "tf64 (2x64)", "acc % (kappa)", "77.9 +- 0.4 (k .71)", "77.0 +- 0.2 (k .69)", "+0.85 (sign consistent 4/4 seeds)",
         "72.8 (k .64)", "75.2", "81-85 (k .77-.83) raw-signal models", "4 pretraining seeds", "results/phase4/gate0/sleep_edf_tf64/finetune.csv; EXP-0020", "yes", "Headline number. Below SOTA by 3-7 with 2.4M params / spectral tokens."],
        ["Sleep staging", "Sleep-EDF-78, subject-disjoint 5-fold", "DE (2x5) — the proposal's token", "acc % (kappa)", "75.5 +- 0.3 (k .672)", "73.1 +- 0.5 (k .645)", "+2.5 (4/4 seeds)",
         "67.9 (k .575)", "69.6", "81-85", "4 pretraining seeds", "results/phase4/gate1/sleep_edf/finetune.csv; EXP-0017 4b, EXP-0021", "yes", "The proposal's pipeline; shows the DE->tf64 gain (+2.3) and that pretraining shrinks as input gets richer."],
        ["Sleep staging", "Sleep-EDF-78, subject-disjoint 5-fold", "raw EEG 200 ms (2x20), 150 tok/epoch", "acc % (kappa)", "75.5 (k .676)", "74.2 (k .658)", "+1.2 (frozen: +14.7)",
         "—", "—", "81-85", "1 seed; 10 pretrain + 3 FT epochs", "results/phase4/gate2/sleep_edf_raw/finetune.csv; EXP-0022", "discuss", "Parity with DE, below tf64 at this budget. Single seed — say so, or run 2 more seeds (~1.5 h on a pod)."],
        ["Sleep staging — streaming", "Sleep-EDF-78, same folds, DE tokens", "DE", "acc % online (only past visible) / offline", "75.4 / 75.4 (causal, pretrained)", "73.2 / 73.2 (causal, no pretrain); bidirectional twin 70.4 online / 74.1 offline", "causal beats bidirectional online by +2.8 (no pretrain) / +5.0 (pretrained)",
         "—", "—", "—", "1 seed, 5 folds", "results/phase4/gate3/streaming.csv; EXP-0023", "yes", "The 'defensible claim' of the causal design; 1 token/decision (KV cache) vs 190."],
        ["Seizure detection", "CHB-MIT, 24 patients, leave-one-patient-out", "DE (18x5)", "bal-acc % / ROC-AUC", "78.4 / .863 (Aug seed 42); 78.2 / .867 (Jul, 3 seeds)", "80.2 / .874 (seed 42); 79.1 / .860 (Jul, 3 seeds)", "≈ 0 (within +-12 per-patient sd)",
         "72.4 / .806", "74.0 / .851", "AUC .91-.99 (cross-patient; verify)", "3 seeds (Jul) + 1 (Aug)", "results/phase3/f17/f17_ft_seed*.csv; results/phase4/gate1/chbmit/finetune.csv; EXP-0017 4c, EXP-0021", "yes", "Architecture beats raw-DE by ~+6 bal-acc / +.06 AUC; pretraining adds nothing. tf64 gives no headroom here (linear 72.4 = 72.4)."],
        "Frozen-probe settings (never fine-tuned) — report as scope / failure cases, not headlines",
        ["Emotion, smoothed", "SEED-IV de_LDS, PC-SSL subject-dependent folds", "DE (62x5)", "acc %", "59.3", "56.9", "+2.4", "62.8", "(nonlinear <= linear, F10)", "84.5 published -> 44.7 clean (leakage)", "3 seeds", "results/phase3/parity; EXP-0016", "discuss", "Pretrained = random 256-d projection (dim-matched -0.2). Below raw features."],
        ["Emotion, un-smoothed", "SEED-IV de_movingAve, same folds", "DE (62x5)", "acc %", "51.7", "40.7", "+11.0 (frozen only)", "55.3", "—", "—", "3 seeds", "results/phase3/parity; EXP-0016", "discuss", "The 'smoothing flip'; still below raw features."],
        ["Motor imagery", "BCI-IV-2a, session hold-out, 9 subj", "DE (22x5)", "acc %", "41.7", "43.5", "-1.8", "51.1", "—", "~70-85 (raw-signal)", "3 seeds", "results/phase3/parity/f16*; EXP-0014/0016", "discuss", "Encoder hurts (dim-matched -9.3). Trial-constant labels."],
        ["Seizure prediction (pre-ictal vs interictal)", "CHB-MIT, patient-specific leave-one-seizure-out, 21 patients / 140 events", "DE (18x5)", "bal-acc % / AUC", "66.8 / .769", "71.3 / .793", "-4.6 / -.024", "65.2 / .710", "—", "—", "1 seed", "FINAL_REPORT §5; EXP-0018 (CSV on pod only)", "no / appendix", "Ran as the falsification test of the 'objective misalignment' mechanism; random-init won. Not comparable to detection."],
        "Reference points",
        ["Phase 1: univariate TimesFM", "SEED-IV/V, same folds", "each (channel,band) trace as a scalar series", "acc %", "20-28 (chance)", "—", "—", "51.4 / 62.8", "—", "—", "—", "docs/PHASE1.md", "yes (one line)", "Why structured patching exists: RevIN + flattening destroy the signal. Structured: 46-61."],
        ["Leakage audit of PC-SSL", "SEED-V / SEED-IV, PC-SSL code", "DE", "acc %", "published 92.4 / 84.5; leaky re-run 65.8 / 70.4; clean trial-disjoint 39.8 / 44.7", "—", "—", "51.4 / 62.8", "—", "—", "—", "results/phase2/followup/f12; EXP-0008", "yes (one paragraph)", "~80% temporal-neighbour leakage. A replication of Brookshire et al. 2024 — cite it."],
    ]
    write_sheet(ws, headers, rows, [22, 30, 22, 16, 26, 26, 22, 16, 16, 26, 16, 40, 14, 50], include_col=12)

    # ------------------------------------------------------------------ 3. Ablations
    ws = wb.create_sheet("Ablations")
    headers = ["#", "Ablation / control", "Question it answers", "Setting", "Arms and numbers", "Delta / effect", "Seeds", "Verdict", "Source", "Include in paper?", "Notes"]
    rows = [
        "A. Does pretraining help? (the contribution under test)",
        [1, "Pretrained vs matched random-init, FINE-TUNED", "Is predictive-coding pretraining worth anything under the protocol competitors use?", "Sleep DE / tf64 / raw / per-electrode; seizure DE",
         "sleep DE 75.5 vs 73.1; tf64 77.9 vs 77.0; raw 75.5 vs 74.2; per-electrode 76.4 vs 74.6; seizure 78.4/.863 vs 80.2/.874 (Jul 3-seed 78.2 vs 79.1)",
         "+2.5 / +0.85 / +1.2 / +1.8 on sleep; ≈ 0 on seizure", "4 / 4 / 1 / 1 / 3+1", "Small positive on sleep, shrinking as input gets richer; null on seizure", "gate1/sleep_edf/finetune.csv, gate0/sleep_edf_tf64/finetune.csv, gate2/*, gate1/chbmit/finetune.csv, f17_ft_seed*", "yes", "THE central ablation; must be in the main table."],
        [2, "Frozen probe vs fine-tuned (evaluation protocol)", "How much did the frozen-probe protocol inflate pretraining?", "same as 1",
         "frozen gaps: DE +9.8, tf64 +11.2, raw +14.7, seizure +8.1; fine-tuned gaps: +2.5, +0.85, +1.2, ≈0", "inflation 3-12x", "mixed", "Frozen probes measure the poverty of a random-init feature extractor, not the value of pretraining", "EXP-0017 4b-c, EXP-0020-0022", "yes", "Methodological finding; one figure (bars frozen vs fine-tuned)."],
        [3, "Dimension-matched control (raw features -> random 256-d nonlinear projection), frozen", "Is the frozen 'gain' just more dimensions?", "sleep DE/tf64, emotion, MI",
         "sleep DE: rand-proj 69.3 vs PC 72.6 (+3.3); tf64: 73.3 vs 76.75 (+3.5); emotion 60.2 vs 60.0 (-0.2); MI 49.9 vs 40.6 (-9.3); concat raw+PC: +5.4 sleep, -0.3 emotion, -7.3 MI",
         "see arms", "1", "Pretrained encoder adds information only on sleep; on emotion it equals a random map", "results/phase3/diagnose_encoder.csv; gate0/diagnose_encoder_tf64.csv; EXP-0017 §3", "yes", ""],
        [4, "Label fraction under fine-tuning", "Does the advantage grow when labels are scarce (the SSL signature)?", "sleep DE, 1/5/10/100% of training labels",
         "frozen gap 12.7/11.8/10.8/9.8; fine-tuned gap 1.9/2.4/2.4/2.2", "flat fine-tuned", "1", "Label-efficiency signature is a frozen-probe artefact (retracted claim)", "results/phase3/f13/f13_sleep_ft_labelcurve.csv; EXP-0017 4d", "discuss", "Mention as part of #2 or drop; do NOT claim 20x/100x."],
        "B. Which pretraining objective?",
        [5, "Input-space PC vs latent-target PC (JEPA/data2vec style)", "Is input-space MSE the reason pretraining fails (R2)?", "sleep DE (4 seeds), seizure, tf64, raw",
         "sleep DE: input 75.5 vs latent 73.7 vs rand 73.1; seizure 78.4 vs 79.7 vs 80.2; tf64 77.9 vs 77.3 vs 77.0; raw 75.5 vs 74.7 vs 74.2",
         "latent -1.9 vs input-PC on sleep DE; ≈ random elsewhere", "4/1/4/1", "Latent targets do not help; objective degenerates into smoothness on per-epoch corpora (skill <= 0)", "gate1/*, EXP-0021", "yes (short)", ""],
        [6, "Latent objective variants", "Any latent variant that works?", "sleep DE, seed 42, fine-tuned",
         "delta targets 73.9; cosine/no-norm 72.7 (collapses); variance term 73.2; EMA .99 74.0; p_out 4 74.0 (input-PC 75.4, rand 73.2)", "all at random-init level", "1", "No", "gate1/sleep_edf/finetune.csv (variant rows); EXP-0021", "appendix", ""],
        [7, "Masked reconstruction vs forecasting (F9)", "Is it the temporal objective?", "emotion smoothed / un-smoothed, frozen",
         "both objectives null on smoothed, both ~+9 over random on un-smoothed", "no difference", "1", "Not the specific objective", "results/phase2/followup/f9; EXP-0011", "appendix", ""],
        [8, "Pretext diagnostics", "Is the pretext learned, and does pretext skill predict transfer?", "all datasets",
         "input-space model/persistence: sleep .60, seizure .57, emo-raw .44, MI .35, emo-smooth 27.8 (degenerate); latent skill vs best trivial: sleep -0.2, seizure -0.6, tf64 ~0, raw +0.36, emotion +0.4/+0.6, MI +0.6",
         "skill anti-correlates with transfer (both objectives)", "1-3", "Pretext learned but misaligned; mechanism's falsifiable prediction (seizure prediction) failed", "results/phase3/diagnose_pretext.csv; gate1/diagnose_pretext_latent.csv; EXP-0017 4e, 0018, 0021", "discuss", "Scientifically the most interesting part; fits Option B/C."],
        "C. Which input tokens?",
        [9, "DE vs tf64 vs raw tokens (architecture, no pretraining)", "What does the input representation buy the architecture?", "sleep, fine-tuned",
         "random-init: DE 73.1, tf64 77.0, raw 74.2; linear ceilings: DE 67.9, tf64 72.8", "tf64 +3.9 over DE for the same architecture", "4/4/1", "tf64 is the better front end on sleep; none on seizure (linear 72.4 = 72.4)", "gate0/saturation_*.csv; finetune csvs; EXP-0020", "yes", "Gate 0."],
        [10, "Linear-saturation test (logreg vs MLP vs HGB)", "Is there nonlinear headroom in the features at all?", "sleep + seizure, DE + tf64",
         "sleep DE 67.9/67.2/69.6; tf64 72.8/71.9/75.2; seizure DE 72.4/69.5/74.0 (AUC .807/.824/.851); tf64 72.4/69.9/72.6",
         "headroom +1.7 -> +2.3 sleep; +1.6 -> +0.2 seizure", "5 folds / 24 pat", "Little headroom on DE; tf64 adds some on sleep only", "gate0/saturation_*; EXP-0010 (emotion), EXP-0020", "yes (one table)", ""],
        [11, "Structured multi-channel tokens vs per-electrode decomposition (BrainGPT-style)", "Is spatial structure in the token what matters?", "sleep raw tokens, 2 channels",
         "structured 75.5/74.7/74.2 vs per-electrode 76.4/75.5/74.6 (PC/latent/rand, fine-tuned); frozen 70.8 vs 72.7", "per-electrode >= structured by ~1", "1", "Not supported on 2-channel sleep; needs a many-channel corpus", "gate2/*; EXP-0022", "discuss", "Honest limitation; do NOT claim structured > per-electrode."],
        [12, "Smoothing flip (LDS-smoothed vs un-smoothed DE)", "Does smoothing hide the dynamics PC needs?", "SEED-IV emotion, frozen, same trials",
         "within-trial variance 0.08% vs 17.6%; PC - rand +2.4 vs +11.0", "+8.6 from removing smoothing", "3", "Pretraining needs dynamics in the tokens — but still below raw features", "EXP-0001, EXP-0016", "discuss", ""],
        [13, "Context length / horizon (F5) and scale (F6)", "Do longer context and multi-horizon help?", "un-smoothed SEED-IV, frozen GRU readout",
         "PC-rand gap grows with p_in (5.3 -> 18.1 at p_in 8/p_out 16); scale-stable 1M-15M (5 -> 13 -> 13)", "—", "1", "Only on un-smoothed DE; frozen readout", "EXP-0005, EXP-0006", "appendix", ""],
        "D. Architecture controls",
        [14, "Causal vs bidirectional twin, offline vs online", "What does causality cost / buy?", "sleep DE, fine-tuned, same folds",
         "causal rand 73.2/73.2; bidir rand 74.1/70.4; causal latent 74.1/74.1; causal PC 75.4/75.4", "offline -0.9; online +2.8 (+5.0 pretrained); 1 vs 190 tokens per decision", "1", "Causal wins under a streaming constraint", "gate3/streaming.csv; EXP-0023", "yes", "Gate 3."],
        [15, "Per-series instance norm (RevIN) control", "Why Phase 1 failed / why no instance norm", "emotion raw DE -> logreg",
         "SEED-V 51.4 -> 18.0; SEED-IV 62.8 -> 26.8 (chance 20/25)", "collapse to chance", "—", "Instance norm deletes the discriminative absolute level", "docs/PHASE2.md E3.2", "yes (one line)", ""],
        [16, "Frozen random stack vs frozen TimesFM stack (F3); matched MLP head (F4)", "Do TimesFM weights transfer? Is the head the limiter?", "emotion, frozen",
         "random 1280x20 stack 47.3/58.9 vs TimesFM 43.3/60.4; MLP head stays in the 41-55 band", "none", "1", "TimesFM weights add nothing; head not the lever", "EXP-0003, EXP-0004", "appendix", ""],
        [17, "Order-shuffle control", "Is the sleep gain temporal?", "sleep frozen; emotion/MI",
         "sleep PC 72.6 -> 67.4 (= raw-DE level), raw unchanged; on trial-constant-label tasks shuffling HELPS (+7.4 emotion) -> uninterpretable there", "-5.2 on sleep", "1", "Frozen sleep gain is temporal; control valid only for per-epoch labels", "EXP-0009 4e, EXP-0016", "appendix", ""],
        [18, "Data-only predictability (tau) vs gain", "Can a raw-data score predict the SSL gain?", "5 settings",
         "MI has the highest tau (0.27) and no gain; sleep the lowest among positives and the largest frozen gain", "no correlation", "—", "No (kills the 'temporal learnability' framing)", "results/phase3/temporal_structure.csv; EXP-0016", "no", ""],
    ]
    write_sheet(ws, headers, rows, [5, 34, 34, 26, 60, 26, 12, 40, 38, 14, 40], include_col=9)

    # ------------------------------------------------------------------ 4. Do not claim
    ws = wb.create_sheet("Do not claim")
    headers = ["Statement", "Why it is out", "Evidence", "What we say instead"]
    rows = [
        ["Pretraining gives +9.8 (sleep) / +14.5 / +8.1 (seizure)", "frozen-probe numbers", "fine-tuned: +2.5 / ≈0", "report fine-tuned; frozen once as the inflation example"],
        ["20x / 100x label efficiency", "frozen-probe artefact", "fine-tuned gap flat +1.9..+2.4 at 1-100% labels", "drop"],
        ["PC helps in proportion to temporal structure", "MI has the highest predictability and no gain; seizure prediction falsified the refined version", "EXP-0016, EXP-0018", "drop"],
        ["The objective-misalignment mechanism explains the failure", "its falsifiable prediction (seizure prediction) failed", "rand .793 > PC .769 AUC", "report the diagnostic (pretext skill anti-correlates with transfer) without asserting the mechanism"],
        ["Structured spatial patching beats per-electrode decomposition", "2-channel sleep shows the opposite", "76.4 vs 75.5", "limitation: needs a many-channel corpus"],
        ["A foundation model / cross-task transfer", "joint multi-dataset pretraining and zero-shot transfer were never run", "proposal Phase 3", "future work; consider renaming"],
        ["Beats / matches published SOTA", "below on every task; SOTA figures unverified", "sleep 77.9 vs 81-85; seizure AUC .87 vs .91-.99", "honest positioning table after verifying numbers; emphasise 2.4M params, spectral tokens, streaming"],
        ["The leakage audit is a new finding", "Brookshire et al. 2024 documented EEG segment-split leakage with similar numbers", "ICML_proposal.md §B", "replication + mechanism, cite Brookshire"],
        ["Latent-target / raw-EEG / richer tokens rescue pretraining", "Gates 0-2", "+0.85 / +1.2 / latent -1.9", "state as tested and negative"],
        ["Seizure prediction is a result", "single seed, patient-specific, run only as a falsification test", "EXP-0018", "appendix at most"],
    ]
    write_sheet(ws, headers, rows, [44, 44, 40, 50])

    wb.save(OUT)
    print("wrote", OUT)


if __name__ == "__main__":
    main()
